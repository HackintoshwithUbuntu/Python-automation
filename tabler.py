from openpyxl import *
from datetime import *

import openpyxl

wb = load_workbook("C:/Users/muzam/Downloads/2021c.xlsx")
names = wb["times"]
table = wb["finals"]

dates = []

"""
for cell in names["A"][1:]:
    val = cell.value.strftime("%m/%d/%Y")
    if val not in dates:
        dates.append(val)
print(dates)
"""
done = False
green = openpyxl.styles.colors.Color(rgb='FF00FF00')
my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=green)

for cell in names["B"][1:]:
    done = False
    crow = cell.row
    ccol = cell.column
    fname = names.cell(row = crow, column = 2).value
    lname = names.cell(row = crow, column = 3).value
    for ncell in table["A"]:
            ncrow = ncell.row
            nfname = ncell.value
            nlname = table.cell(row = ncrow, column = 2).value
            if nfname == fname and nlname == lname:
                table.cell(row = ncrow, column = 1).value = fname
                table.cell(row = ncrow, column = 2).value = lname
                table.cell(row = ncrow, column = 3).fill = my_fill
                done = True
    if not done:
        lastr = int(table.max_row) + 1
        table.cell(row = lastr, column = 1).value = fname
        table.cell(row = lastr, column = 2).value = lname
        table.cell(row = lastr, column = 3).fill = my_fill


wb.save("C:/Users/muzam/Downloads/2021c.xlsx")