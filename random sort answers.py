from openpyxl import *
import random as rand
from random import *

wb = load_workbook("questions2.xlsx")
ws = wb["Table 1"]

valuestore = []
locstore = []

for scell in ws["A"]:
    #print(cell.value)
    result = str(scell.value)
    if(result[0] == '-'):
        valuestore.append(result)
        locstore.append(int(scell.row))
    if(len(valuestore) == 3):
        rand.shuffle(valuestore)
        for i in range(3):
            mycell = ws.cell(row = locstore[i], column = 1)
            mycell.value = valuestore[i]
        valuestore = []
        locstore = []
wb.save("out2.xlsx")