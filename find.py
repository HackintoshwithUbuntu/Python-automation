from openpyxl import *
import random as rand
from random import *


def work(wbname, savename, namesheet, namerow, sortedsheet):
    wb = load_workbook(wbname)
    ws = wb[sortedsheet]
    nws = wb[namesheet]
    rcalls = []

    for cell in ws["A"]:
        crow = cell.row
        fname = cell.value
        lname = ws.cell(row = crow, column = 2).value
        for ncell in nws["B"]:
            ncrow = ncell.row
            nfname = ncell.value
            nlname = nws.cell(row = ncrow, column = 3).value
            if nfname == fname and nlname == lname:
                rcall = nws.cell(row = ncrow, column = 5).value
                rcalls.append(rcall)
        for i in range(len(rcalls)):
            ecell = ws.cell(row = crow, column = 11 + i)
            ecell.value = rcalls[i]
        rcalls = []

    wb.save(savename)

work("counted.xlsx", "final.xlsx", "names", "B", "do")
